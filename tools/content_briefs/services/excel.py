from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ── Palette ───────────────────────────────────────────────────────────────────
NAV_FILL    = PatternFill('solid', fgColor='02054F')  # navy — main header
SCOPE_FILL  = PatternFill('solid', fgColor='1A3A8F')  # mid blue — scoped cols
LOCALE_FILL = PatternFill('solid', fgColor='1A6B3A')  # green — locale cols
EXTRA_FILL  = PatternFill('solid', fgColor='7A4400')  # amber — extra cols
REF_FILL    = PatternFill('solid', fgColor='1A5276')  # teal-navy — reference cols
HAS_FILL    = PatternFill('solid', fgColor='E8E8E8')  # light grey — has value
NEED_FILL   = PatternFill('solid', fgColor='FFF3CC')  # yellow — needs writing
MISS_FILL   = PatternFill('solid', fgColor='FFE0E0')  # red tint — SKU not found
GOLD_FONT   = Font(bold=True, color='F2C400', size=10)
WHITE_FONT  = Font(bold=True, color='FFFFFF', size=10)
BODY_FONT   = Font(size=10)
THIN_BORDER = Border(
    left=Side(style='thin', color='C0C0C0'),
    right=Side(style='thin', color='C0C0C0'),
    bottom=Side(style='thin', color='C0C0C0'),
)


def _hdr(ws, row, col, text, fill, font):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = fill
    c.font = font
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = THIN_BORDER
    return c


def generate_content_brief(rows, output_path, scoped_attrs, locale_attrs, extra_attrs,
                            brief_name='', notes=None, reference_attrs=None):
    wb = Workbook()

    # ── Sheet 1: Brief data ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Content Brief'

    # Build column list
    col_defs = []  # (label, type, key, attr)
    col_defs.append(('SKU', 'meta', None, None))
    col_defs.append(('Found in Akeneo', 'meta', None, None))

    for scope, attrs in scoped_attrs.items():
        for attr in attrs:
            col_defs.append((f'[{scope}]\n{attr}', 'scoped', scope, attr))

    for locale, attrs in locale_attrs.items():
        for attr in attrs:
            col_defs.append((f'[{locale}]\n{attr}', 'locale', locale, attr))

    for attr in extra_attrs:
        col_defs.append((attr, 'extra', None, attr))

    for attr in (reference_attrs or []):
        col_defs.append((f'REF\n{attr}', 'reference', None, attr))

    # Row 1: headers
    for col_idx, (label, col_type, _, _) in enumerate(col_defs, 1):
        if col_type == 'meta':
            fill, font = NAV_FILL, GOLD_FONT
        elif col_type == 'scoped':
            fill, font = SCOPE_FILL, WHITE_FONT
        elif col_type == 'locale':
            fill, font = LOCALE_FILL, WHITE_FONT
        elif col_type == 'reference':
            fill, font = REF_FILL, WHITE_FONT
        else:
            fill, font = EXTRA_FILL, WHITE_FONT
        _hdr(ws, 1, col_idx, label, fill, font)

    ws.row_dimensions[1].height = 48

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        not_found = not row['found']
        row_fill = MISS_FILL if not_found else None

        sku_cell = ws.cell(row=row_idx, column=1, value=row['sku'])
        sku_cell.font = Font(bold=True, size=10)
        sku_cell.alignment = Alignment(vertical='top')
        if row_fill:
            sku_cell.fill = row_fill

        found_cell = ws.cell(row=row_idx, column=2, value='No — SKU not found' if not_found else 'Yes')
        found_cell.font = Font(color='CC0000' if not_found else '2E7D32', size=10)
        found_cell.alignment = Alignment(vertical='top', horizontal='center')
        if row_fill:
            found_cell.fill = row_fill

        for col_idx, (_, col_type, key, attr) in enumerate(col_defs[2:], 3):
            if col_type == 'scoped':
                value = row['values']['scoped'].get(key, {}).get(attr, '')
            elif col_type == 'locale':
                value = row['values']['locale'].get(key, {}).get(attr, '')
            elif col_type == 'reference':
                value = row['values'].get('reference', {}).get(attr, '')
            else:
                value = row['values']['extra'].get(attr, '')

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = THIN_BORDER
            if not_found:
                cell.fill = MISS_FILL
            elif col_type == 'reference':
                cell.fill = HAS_FILL  # reference cols are always grey — context only
            else:
                cell.fill = HAS_FILL if value else NEED_FILL

    # Column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18
    for col_idx in range(3, len(col_defs) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 42

    ws.freeze_panes = 'A2'

    # ── Sheet 2: Config ──────────────────────────────────────────────────────
    cfg = wb.create_sheet('Brief Config')
    cfg.column_dimensions['A'].width = 22
    cfg.column_dimensions['B'].width = 60

    def cfg_row(r, key, val):
        a = cfg.cell(row=r, column=1, value=key)
        a.font = Font(bold=True, size=10)
        a.fill = NAV_FILL
        a.font = GOLD_FONT
        a.alignment = Alignment(vertical='top')
        b = cfg.cell(row=r, column=2, value=val)
        b.font = BODY_FONT
        b.alignment = Alignment(wrap_text=True, vertical='top')
        cfg.row_dimensions[r].height = max(15, len(str(val)) // 3)

    r = 1
    cfg_row(r, 'Brief Name', brief_name); r += 1
    cfg_row(r, 'Generated', datetime.now().strftime('%Y-%m-%d %H:%M')); r += 1
    cfg_row(r, 'SKU Count', len(rows)); r += 1
    cfg_row(r, 'Notes', notes or ''); r += 1
    r += 1

    cfg_row(r, 'Scoped Attributes', ''); r += 1
    for scope, attrs in scoped_attrs.items():
        cfg_row(r, f'  {scope}', ', '.join(attrs)); r += 1
    r += 1

    cfg_row(r, 'Localised Attributes', ''); r += 1
    for locale, attrs in locale_attrs.items():
        cfg_row(r, f'  {locale}', ', '.join(attrs)); r += 1
    r += 1

    cfg_row(r, 'Extra Attributes', ', '.join(extra_attrs) if extra_attrs else '—'); r += 1
    r += 1
    cfg_row(r, 'Reference Columns', ', '.join(reference_attrs) if reference_attrs else '—')

    # ── Sheet 3: Key ────────────────────────────────────────────────────────
    key_ws = wb.create_sheet('Key')
    key_ws.column_dimensions['A'].width = 25
    key_ws.column_dimensions['B'].width = 50

    key_rows = [
        ('Column colour', 'Meaning'),
        ('Navy header', 'SKU / metadata columns'),
        ('Blue header', 'Scoped attribute (per sales channel)'),
        ('Green header', 'Localised attribute (per language)'),
        ('Amber header', 'Extra / global attribute'),
        ('Teal header', 'Reference column — context only, not for writing'),
        ('', ''),
        ('Cell colour', 'Meaning'),
        ('Grey cell', 'Has an existing value in Akeneo — review and update if needed'),
        ('Yellow cell', 'No value in Akeneo — needs writing'),
        ('Pink/red cell', 'SKU not found in Akeneo'),
    ]
    fills_key = [NAV_FILL, NAV_FILL, SCOPE_FILL, LOCALE_FILL, EXTRA_FILL, REF_FILL,
                 None, NAV_FILL, HAS_FILL, NEED_FILL, MISS_FILL]

    for i, ((k, v), fill) in enumerate(zip(key_rows, fills_key), 1):
        ca = key_ws.cell(row=i, column=1, value=k)
        cb = key_ws.cell(row=i, column=2, value=v)
        ca.font = GOLD_FONT if fill in (NAV_FILL, SCOPE_FILL, LOCALE_FILL, EXTRA_FILL) else BODY_FONT
        cb.font = BODY_FONT
        if fill:
            ca.fill = fill
            cb.fill = fill

    wb.save(output_path)
