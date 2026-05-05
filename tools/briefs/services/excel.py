from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side


def generate_brief_spreadsheet(brief_data, output_path, companion_models=None,
                                scene_mode='new', scene_ref='', recolour_notes='',
                                include_sp1=False):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Brief'

    headers = [
        'Nuie Sku', 'TITLE', 'COLOUR', 'CUTOUT LINK', 'CATEGORY', 'FAMILY',
        'Front Lifestyle', 'file name', 'Top Lifestyle', 'file name', 'Side Lifestyle', 'file name',
        'CC1', 'CC2', 'CC3', 'CC4', 'CC5', 'CC6',
        'FULL PRODUCT CODE', '3D Models', 'Model Files', 'Scene Count', 'Scene Vibe / Mood',
        'Companion Models (Non-Essential)', 'Scene Mode', 'Scene Reference', 'Recolour Notes',
        *(['Splash Shot (sp1)', 'sp1 file name'] if include_sp1 else [])
    ]

    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    header_font = Font(bold=True)
    thin_border = Border(bottom=Side(style='thin'))

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    missing_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    companion_str = '; '.join(companion_models) if companion_models else ''

    for row_num, item in enumerate(brief_data, 2):
        ws.cell(row=row_num, column=1, value=item.get('sku', ''))
        ws.cell(row=row_num, column=2, value=item.get('title', ''))
        ws.cell(row=row_num, column=3, value=item.get('colour', ''))
        ws.cell(row=row_num, column=4, value=item.get('cutout_link', ''))
        ws.cell(row=row_num, column=5, value=item.get('category', ''))
        ws.cell(row=row_num, column=6, value=item.get('family', ''))
        ws.cell(row=row_num, column=7, value='yes' if item.get('front_lifestyle') else '')
        ws.cell(row=row_num, column=8, value=item.get('front_lifestyle_file', ''))
        ws.cell(row=row_num, column=9, value='yes' if item.get('top_lifestyle') else '')
        ws.cell(row=row_num, column=10, value=item.get('top_lifestyle_file', ''))
        ws.cell(row=row_num, column=11, value='yes' if item.get('side_lifestyle') else '')
        ws.cell(row=row_num, column=12, value=item.get('side_lifestyle_file', ''))
        ws.cell(row=row_num, column=13, value=item.get('cc1', ''))
        ws.cell(row=row_num, column=14, value=item.get('cc2', ''))
        ws.cell(row=row_num, column=15, value=item.get('cc3', ''))
        ws.cell(row=row_num, column=16, value=item.get('cc4', ''))
        ws.cell(row=row_num, column=17, value=item.get('cc5', ''))
        ws.cell(row=row_num, column=18, value=item.get('cc6', ''))
        ws.cell(row=row_num, column=19, value=item.get('full_product_code', ''))
        ws.cell(row=row_num, column=20, value=item.get('models', ''))
        ws.cell(row=row_num, column=21, value=item.get('model_files', ''))
        ws.cell(row=row_num, column=22, value=item.get('scene_count', 3))
        ws.cell(row=row_num, column=23, value=item.get('scene_vibe', ''))
        if row_num == 2:
            if companion_str:
                ws.cell(row=row_num, column=24, value=companion_str)
            if scene_mode != 'new':
                ws.cell(row=row_num, column=25, value=scene_mode)
                ws.cell(row=row_num, column=26, value=scene_ref)
                if recolour_notes:
                    ws.cell(row=row_num, column=27, value=recolour_notes)
        if include_sp1:
            ws.cell(row=row_num, column=28, value='yes' if item.get('splash_shot') else '')
            ws.cell(row=row_num, column=29, value=item.get('sp1_file', ''))
        if not item.get('models'):
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = missing_fill

    for col_num, header in enumerate(headers, 1):
        max_length = len(header)
        for row in range(2, len(brief_data) + 2):
            cell_value = ws.cell(row=row, column=col_num).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = min(max_length + 2, 50)

    wb.save(output_path)
    return output_path
