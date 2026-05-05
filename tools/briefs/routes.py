import os
import re
import io
import csv
import json
import zipfile
import shutil
import sqlite3
from datetime import datetime, timedelta

import requests
from flask import render_template, request, redirect, url_for, session, flash, jsonify, send_file, abort, g, Response
from werkzeug.utils import secure_filename

from tools.briefs import briefs_bp
from shared.auth import tool_access_required, get_current_user

from tools.briefs.services.akeneo import AkeneoClient
from tools.briefs.services.models import ModelLookup
from tools.briefs.services.excel import generate_brief_spreadsheet
from tools.briefs.services.scaleflex import ScaleflexClient
from tools.briefs.services.notifications import notify_brief_assigned, notify_priority_changed

# ── Config ───────────────────────────────────────────────────────────────────

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'data')
DB_PATH  = os.path.join(DATA_DIR, 'briefs.db')
BRIEFS_OUTPUT_DIR = os.path.join(BASE_DIR, 'briefs_output')

os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(BRIEFS_OUTPUT_DIR, exist_ok=True)
os.makedirs(os.path.join(BASE_DIR, 'uploads'), exist_ok=True)

CONFIG = {
    'output_base': BRIEFS_OUTPUT_DIR,
    'database': DB_PATH,
    's3': {
        'bucket': os.environ.get('S3_BUCKET', '3d-models-lifestyle-briefs'),
        'region': os.environ.get('AWS_REGION', 'eu-west-2'),
        'access_key': os.environ.get('AWS_ACCESS_KEY_ID', ''),
        'secret_key': os.environ.get('AWS_SECRET_ACCESS_KEY', ''),
        'models_prefix': '4-3D/',
        'mapping_key': 'FINAL_master_mapping.csv'
    },
    'akeneo': {
        'base_url': os.environ.get('AKENEO_URL', ''),
        'client_id': os.environ.get('AKENEO_CLIENT_ID', ''),
        'client_secret': os.environ.get('AKENEO_CLIENT_SECRET', ''),
        'username': os.environ.get('AKENEO_USERNAME', ''),
        'password': os.environ.get('AKENEO_PASSWORD', '')
    },
    'local': {
        'enabled': os.environ.get('USE_LOCAL', 'true').lower() == 'true',
        'master_mapping': os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
            '..', 'lifestyle operations', 'FINAL_master_mapping.csv'
        ),
        'models_base': os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
            '..', 'lifestyle operations', '4-3D'
        ),
    }
}

akeneo_client = AkeneoClient(CONFIG['akeneo'])
model_lookup   = ModelLookup(CONFIG)
scaleflex_client = ScaleflexClient()

ROOM_TYPES = ['Bathroom', 'Kitchen', 'En-Suite', 'Cloakroom', 'WC', 'Utility', 'Bedroom', 'Other']

# ── DB ───────────────────────────────────────────────────────────────────────

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()

    conn.execute('''CREATE TABLE IF NOT EXISTS briefs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        name TEXT NOT NULL,
        sku_count INTEGER,
        models_found INTEGER,
        models_missing INTEGER,
        output_path TEXT,
        brief_type TEXT DEFAULT 'lifestyle',
        status TEXT DEFAULT 'pending',
        priority TEXT DEFAULT 'normal',
        assigned_to INTEGER,
        deadline TEXT,
        downloaded_at TEXT,
        downloaded_by INTEGER,
        scene_count INTEGER DEFAULT 3,
        scene_vibe TEXT,
        sku_overrides TEXT,
        sku_list TEXT,
        companion_models TEXT,
        missing_model_skus TEXT,
        deadline_days INTEGER,
        include_sp1 INTEGER DEFAULT 0,
        scene_mode TEXT DEFAULT 'new',
        selected_scenes TEXT,
        recolour_notes TEXT,
        completed_at TEXT,
        amendment_count INTEGER DEFAULT 0,
        is_maam INTEGER DEFAULT 0,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')

    conn.execute('''CREATE TABLE IF NOT EXISTS model_mappings (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        sku TEXT UNIQUE NOT NULL,
        models TEXT NOT NULL,
        model_count INTEGER DEFAULT 1,
        source TEXT DEFAULT 'manual',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')

    conn.execute('''CREATE TABLE IF NOT EXISTS excluded_skus (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        sku TEXT UNIQUE NOT NULL,
        reason TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')

    conn.execute('''CREATE TABLE IF NOT EXISTS scenes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        description TEXT,
        room_type TEXT,
        created_by INTEGER NOT NULL,
        preview_s3_key TEXT,
        tags TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')

    conn.execute('''CREATE TABLE IF NOT EXISTS scene_files (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        scene_id INTEGER NOT NULL,
        filename TEXT NOT NULL,
        s3_key TEXT NOT NULL,
        file_size INTEGER,
        uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (scene_id) REFERENCES scenes(id) ON DELETE CASCADE
    )''')

    conn.execute('''CREATE TABLE IF NOT EXISTS app_settings (
        key TEXT PRIMARY KEY,
        value TEXT NOT NULL,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')
    conn.execute("INSERT OR IGNORE INTO app_settings (key, value) VALUES ('hourly_rate', '8')")

    conn.execute('''CREATE TABLE IF NOT EXISTS brief_amendments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        brief_id INTEGER NOT NULL,
        sent_back_at TEXT NOT NULL,
        sent_back_by INTEGER,
        reason TEXT,
        completed_at TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')

    conn.execute('''CREATE TABLE IF NOT EXISTS uploads (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        uploaded_by INTEGER NOT NULL,
        original_filename TEXT NOT NULL,
        s3_key TEXT NOT NULL,
        detected_sku TEXT,
        confirmed_sku TEXT,
        image_type TEXT,
        brief_id INTEGER,
        status TEXT DEFAULT 'pending',
        reviewer_id INTEGER,
        review_comment TEXT,
        reviewed_at TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')

    conn.commit()
    conn.close()


init_db()

# ── Helpers ──────────────────────────────────────────────────────────────────

def is_admin():
    user = g.get('user') or get_current_user()
    return user and user['role'] == 'admin'


def detect_sku_from_filename(filename):
    name = os.path.splitext(filename)[0]
    image_type = None
    type_match = re.search(r'_(ls\d|sp\d)', name, re.IGNORECASE)
    if type_match:
        image_type = type_match.group(1).lower()
        sku = name[:type_match.start()]
    else:
        sku = name
    sku = sku.strip(' _-')
    return sku if sku else None, image_type


def _get_s3_client():
    import boto3
    s3_cfg = CONFIG['s3']
    return boto3.client('s3',
        region_name=s3_cfg.get('region', 'eu-west-2'),
        aws_access_key_id=s3_cfg.get('access_key') or None,
        aws_secret_access_key=s3_cfg.get('secret_key') or None
    )


def upload_to_s3(file_obj, s3_key, content_type=None):
    s3 = _get_s3_client()
    extra = {'ContentType': content_type} if content_type else {}
    s3.upload_fileobj(file_obj, CONFIG['s3']['bucket'], s3_key, ExtraArgs=extra)


def delete_from_s3(s3_key):
    s3 = _get_s3_client()
    s3.delete_object(Bucket=CONFIG['s3']['bucket'], Key=s3_key)


def get_s3_presigned_url(s3_key, expires=3600):
    s3 = _get_s3_client()
    return s3.generate_presigned_url('get_object',
        Params={'Bucket': CONFIG['s3']['bucket'], 'Key': s3_key}, ExpiresIn=expires)


# ── Dashboard ─────────────────────────────────────────────────────────────────

@briefs_bp.route('/')
@tool_access_required('briefs')
def index():
    db = get_db()
    user = g.user

    if user['role'] == 'admin':
        all_briefs = db.execute('''
            SELECT b.*
            FROM briefs b
            ORDER BY
                CASE b.priority WHEN 'urgent' THEN 1 WHEN 'high' THEN 2 WHEN 'normal' THEN 3 WHEN 'low' THEN 4 END,
                b.created_at DESC
        ''').fetchall()
        new_briefs = [b for b in all_briefs if b['status'] in (None, 'pending', 'assigned')]
        in_progress = [b for b in all_briefs if b['status'] == 'in_progress']
        complete = [b for b in all_briefs if b['status'] == 'complete']
        pending_uploads = db.execute("SELECT COUNT(*) as c FROM uploads WHERE status = 'pending'").fetchone()['c']
        db.close()
        today = datetime.now().strftime('%Y-%m-%d')
        soon = (datetime.now() + timedelta(days=2)).strftime('%Y-%m-%d')
        return render_template('briefs/index.html',
                               new_briefs=new_briefs, in_progress=in_progress,
                               complete=complete, pending_uploads=pending_uploads,
                               today=today, soon=soon)
    else:
        briefs = db.execute(
            'SELECT * FROM briefs WHERE user_id = ? ORDER BY created_at DESC',
            (user['id'],)
        ).fetchall()
        my_uploads = db.execute(
            'SELECT * FROM uploads WHERE uploaded_by = ? ORDER BY created_at DESC LIMIT 20',
            (user['id'],)
        ).fetchall()
        db.close()
        return render_template('briefs/index.html',
                               briefs=briefs, my_uploads=my_uploads)


# ── Generate Brief ───────────────────────────────────────────────────────────

@briefs_bp.route('/generate', methods=['GET', 'POST'])
@tool_access_required('briefs')
def generate():
    if request.method == 'POST':
        brief_name = request.form.get('brief_name', '').strip()
        brief_type = request.form.get('brief_type', 'lifestyle')
        sku_list = request.form.get('sku_list', '')

        if 'sku_file' in request.files:
            f = request.files['sku_file']
            if f and f.filename:
                try:
                    sku_list = f.read().decode('utf-8')
                except UnicodeDecodeError:
                    flash('File must be UTF-8 encoded', 'error')
                    return render_template('briefs/generate.html')

        skus = list(dict.fromkeys([s.strip() for s in sku_list.replace('\r', '').split('\n') if s.strip()]))

        if not brief_name:
            flash('Please enter a brief name', 'error')
            return render_template('briefs/generate.html')
        if not skus:
            flash('Please provide at least one SKU', 'error')
            return render_template('briefs/generate.html')

        try:
            scene_count = int(request.form.get('scene_count', '3'))
        except ValueError:
            scene_count = 3

        scene_vibe = request.form.get('scene_vibe', '')
        try:
            category_brand_overrides = json.loads(request.form.get('category_brand_overrides', '{}') or '{}')
        except json.JSONDecodeError:
            category_brand_overrides = {}
        try:
            companion_models = json.loads(request.form.get('companion_models', '[]') or '[]')
        except json.JSONDecodeError:
            companion_models = []

        deadline = request.form.get('deadline', '')
        deadline_days_raw = request.form.get('deadline_days', '')
        deadline_days = int(deadline_days_raw) if deadline_days_raw else None
        exclude_missing_models = request.form.get('exclude_missing_models') == 'on'
        include_sp1 = request.form.get('include_sp1') == 'on'
        scene_mode = request.form.get('scene_mode', 'new')
        try:
            selected_scenes = json.loads(request.form.get('selected_scenes', '[]') or '[]')
        except json.JSONDecodeError:
            selected_scenes = []
        recolour_notes = request.form.get('recolour_notes', '').strip()

        session['pending_brief'] = {
            'name': brief_name,
            'skus': skus,
            'brief_type': brief_type,
            'scene_count': scene_count,
            'scene_vibe': scene_vibe,
            'category_brand_overrides': category_brand_overrides,
            'companion_models': companion_models,
            'deadline': deadline,
            'deadline_days': deadline_days,
            'exclude_missing_models': exclude_missing_models,
            'include_sp1': include_sp1,
            'scene_mode': scene_mode,
            'selected_scenes': selected_scenes,
            'recolour_notes': recolour_notes
        }
        return redirect(url_for('briefs.processing'))

    db = get_db()
    scenes = db.execute('SELECT id, name, room_type FROM scenes ORDER BY name').fetchall()
    db.close()
    return render_template('briefs/generate.html', scenes=scenes)


@briefs_bp.route('/processing')
@tool_access_required('briefs')
def processing():
    if 'pending_brief' not in session:
        return redirect(url_for('briefs.generate'))
    return render_template('briefs/processing.html',
                           brief_name=session['pending_brief']['name'],
                           sku_count=len(session['pending_brief']['skus']))


@briefs_bp.route('/api/process', methods=['POST'])
@tool_access_required('briefs')
def api_process():
    if 'pending_brief' not in session:
        return jsonify({'error': 'No pending brief'}), 400

    try:
        pb = session['pending_brief']
        name = pb['name']
        skus = pb['skus']
        brief_type = pb.get('brief_type', 'lifestyle')
        scene_count = pb.get('scene_count', 3)
        scene_vibe = pb.get('scene_vibe', '')
        category_brand_overrides = pb.get('category_brand_overrides', {})
        companion_models = pb.get('companion_models', [])
        exclude_missing_models = pb.get('exclude_missing_models', False)
        include_sp1 = pb.get('include_sp1', False)
        scene_mode = pb.get('scene_mode', 'new')
        selected_scenes = pb.get('selected_scenes', [])
        recolour_notes = pb.get('recolour_notes', '')
        include_models = (brief_type == 'lifestyle')

        timestamp = datetime.now().strftime('%Y-%m-%d_%H%M%S')
        output_folder = os.path.join(CONFIG['output_base'], f'{name}_{timestamp}')
        os.makedirs(output_folder, exist_ok=True)

        if include_models:
            os.makedirs(os.path.join(output_folder, '3D_Models'), exist_ok=True)

        products = akeneo_client.get_products(skus)

        brief_data = []
        models_found = models_missing = 0
        excluded_skus_list = []

        for sku in skus:
            product = products.get(sku)
            model_info = model_lookup.get_models_for_sku(sku)
            has_model = model_info['found'] and model_info['files']

            if has_model:
                models_found += 1
                if include_models:
                    model_lookup.copy_model_files(model_info['files'], os.path.join(output_folder, '3D_Models'))
            else:
                models_missing += 1
                if exclude_missing_models and include_models:
                    excluded_skus_list.append(sku)
                    continue

            sku_category = product['categories'][0] if product and product.get('categories') else ''
            sku_brand = (akeneo_client.get_attribute(product, 'brand') or '') if product else ''
            override_key = f'{sku_category}|{sku_brand}'
            cat_brand_override = category_brand_overrides.get(override_key, {})
            sku_scene_count = cat_brand_override.get('scene_count', scene_count)
            sku_scene_vibe = cat_brand_override.get('vibe', scene_vibe)

            brief_data.append({
                'sku': sku,
                'title': akeneo_client.get_attribute(product, 'title') if product else '',
                'colour': akeneo_client.get_attribute(product, 'colour_marketing') if product else '',
                'cutout_link': akeneo_client.get_asset_url(product, 'cutout_1') if product else '',
                'category': sku_category,
                'family': product.get('family', '') if product else '',
                'front_lifestyle': True, 'front_lifestyle_file': f'{sku}_ls1',
                'top_lifestyle': True, 'top_lifestyle_file': f'{sku}_ls2',
                'side_lifestyle': True, 'side_lifestyle_file': f'{sku}_ls3',
                'splash_shot': include_sp1, 'sp1_file': f'{sku}_sp1' if include_sp1 else '',
                'cc1': '', 'cc2': '', 'cc3': '', 'cc4': '', 'cc5': '', 'cc6': '',
                'full_product_code': sku,
                'models': '; '.join(model_info['models']) if model_info['found'] else '',
                'model_files': '; '.join([os.path.basename(f) for f in model_info['files']]),
                'scene_count': sku_scene_count,
                'scene_vibe': sku_scene_vibe
            })

        if include_models and companion_models:
            model_lookup._load_mapping()
            model_lookup._index_model_files()
            for comp_model in companion_models:
                if comp_model in model_lookup.model_files_index:
                    model_lookup.copy_model_files(model_lookup.model_files_index[comp_model],
                                                  os.path.join(output_folder, '3D_Models'))

        scene_ref = ''
        if scene_mode in ('existing', 'recolour', 'mix') and selected_scenes:
            sdb = get_db()
            placeholders = ','.join('?' * len(selected_scenes))
            scene_rows = sdb.execute(f'SELECT name FROM scenes WHERE id IN ({placeholders})', selected_scenes).fetchall()
            sdb.close()
            scene_ref = '; '.join(r['name'] for r in scene_rows)

        excel_path = os.path.join(output_folder, f'{name}.xlsx')
        generate_brief_spreadsheet(brief_data, excel_path, companion_models=companion_models,
                                   scene_mode=scene_mode, scene_ref=scene_ref,
                                   recolour_notes=recolour_notes, include_sp1=include_sp1)

        cutouts_folder = os.path.join(output_folder, 'Cutouts')
        cutouts_downloaded = 0
        for item in brief_data:
            cutout_url = item.get('cutout_link', '')
            if cutout_url:
                try:
                    img_resp = requests.get(cutout_url, timeout=15)
                    if img_resp.status_code == 200:
                        os.makedirs(cutouts_folder, exist_ok=True)
                        ct = img_resp.headers.get('Content-Type', '')
                        ext = '.png' if 'png' in ct else '.jpg'
                        with open(os.path.join(cutouts_folder, f"{item['sku']}_cutout{ext}"), 'wb') as f:
                            f.write(img_resp.content)
                        cutouts_downloaded += 1
                except Exception as e:
                    print(f"Cutout download failed for {item['sku']}: {e}")

        user = g.user
        db = get_db()
        cursor = db.execute(
            '''INSERT INTO briefs (user_id, name, sku_count, models_found, models_missing,
               output_path, brief_type, scene_count, scene_vibe, sku_overrides, sku_list,
               companion_models, missing_model_skus, deadline, deadline_days, status,
               scene_mode, selected_scenes, recolour_notes, include_sp1)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
            (user['id'], name, len(brief_data), models_found, models_missing,
             output_folder, brief_type, scene_count, scene_vibe,
             json.dumps(category_brand_overrides), json.dumps(skus),
             json.dumps(companion_models) if companion_models else None,
             json.dumps(excluded_skus_list) if excluded_skus_list else None,
             pb.get('deadline') or None, pb.get('deadline_days'),
             'pending', scene_mode,
             json.dumps(selected_scenes) if selected_scenes else None,
             recolour_notes or None, 1 if include_sp1 else 0)
        )
        db.commit()
        brief_id = cursor.lastrowid
        db.close()

        session.pop('pending_brief', None)

        return jsonify({
            'success': True,
            'brief_id': brief_id,
            'name': name,
            'sku_count': len(brief_data),
            'models_found': models_found,
            'models_missing': models_missing,
            'output_folder': output_folder,
            'brief_type': brief_type,
            'excluded_skus': excluded_skus_list
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ── Brief Management ─────────────────────────────────────────────────────────

@briefs_bp.route('/download-all/<int:brief_id>')
@tool_access_required('briefs')
def download_all(brief_id):
    db = get_db()
    user = g.user
    if user['role'] == 'admin':
        brief = db.execute('SELECT * FROM briefs WHERE id = ?', (brief_id,)).fetchone()
    else:
        brief = db.execute('SELECT * FROM briefs WHERE id = ? AND user_id = ?', (brief_id, user['id'])).fetchone()

    if brief and brief['output_path'] and os.path.exists(brief['output_path']):
        now_str = datetime.now().strftime('%Y-%m-%d %H:%M')
        if brief['status'] in (None, 'pending', 'assigned'):
            new_deadline = brief['deadline']
            if brief['deadline_days'] and not brief['deadline']:
                new_deadline = (datetime.now() + timedelta(days=brief['deadline_days'])).strftime('%Y-%m-%d')
            db.execute('UPDATE briefs SET downloaded_at = ?, downloaded_by = ?, status = ?, deadline = ? WHERE id = ?',
                       (now_str, user['id'], 'in_progress', new_deadline, brief_id))
        else:
            db.execute('UPDATE briefs SET downloaded_at = ?, downloaded_by = ? WHERE id = ?',
                       (now_str, user['id'], brief_id))
        db.commit()
        db.close()

        memory_file = io.BytesIO()
        with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
            for root, dirs, files in os.walk(brief['output_path']):
                for file in files:
                    file_path = os.path.join(root, file)
                    arc_name = os.path.relpath(file_path, brief['output_path'])
                    zf.write(file_path, arc_name)
        memory_file.seek(0)
        return send_file(memory_file, mimetype='application/zip', as_attachment=True,
                         download_name=f"{brief['name']}_complete.zip")

    db.close()
    flash('Files not found', 'error')
    return redirect(url_for('briefs.index'))


@briefs_bp.route('/download-excluded/<int:brief_id>')
@tool_access_required('briefs')
def download_excluded(brief_id):
    db = get_db()
    brief = db.execute('SELECT id, name, missing_model_skus FROM briefs WHERE id = ?', (brief_id,)).fetchone()
    db.close()
    if not brief or not brief['missing_model_skus']:
        flash('No excluded SKUs for this brief', 'error')
        return redirect(url_for('briefs.index'))
    excluded = json.loads(brief['missing_model_skus'])
    if not excluded:
        flash('No excluded SKUs for this brief', 'error')
        return redirect(url_for('briefs.index'))
    output = io.StringIO()
    csv.writer(output).writerows([['sku']] + [[s] for s in excluded])
    mem = io.BytesIO(output.getvalue().encode('utf-8'))
    return send_file(mem, mimetype='text/csv', as_attachment=True,
                     download_name=f"{brief['name']}_excluded_skus.csv")


@briefs_bp.route('/view-brief/<int:brief_id>')
@tool_access_required('briefs')
def view_brief(brief_id):
    db = get_db()
    brief = db.execute('SELECT * FROM briefs WHERE id = ?', (brief_id,)).fetchone()
    db.close()
    if not brief:
        flash('Brief not found', 'error')
        return redirect(url_for('briefs.index'))

    rows = []
    excel_file = os.path.join(brief['output_path'], f"{brief['name']}.xlsx") if brief['output_path'] else None
    if excel_file and os.path.exists(excel_file):
        from openpyxl import load_workbook
        wb = load_workbook(excel_file, read_only=True)
        ws = wb.active
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx == 0:
                continue
            cells = list(row)
            while len(cells) < 29:
                cells.append(None)
            v = lambda i: cells[i] if cells[i] is not None else ''
            rows.append({
                'sku': v(0), 'title': v(1), 'colour': v(2), 'cutout_link': v(3),
                'category': v(4), 'family': v(5),
                'ls1_file': v(7), 'ls2_file': v(9), 'ls3_file': v(11),
                'models': v(19), 'model_files': v(20),
                'scene_count': v(21), 'scene_vibe': v(22),
                'companions': v(23), 'sp1_file': v(28),
            })
        wb.close()

    models_with = sum(1 for r in rows if r['models'])
    models_without = sum(1 for r in rows if not r['models'])
    return render_template('briefs/view_brief.html', brief=brief, rows=rows,
                           models_with=models_with, models_without=models_without)


@briefs_bp.route('/assign-brief/<int:brief_id>', methods=['POST'])
@tool_access_required('briefs')
def assign_brief(brief_id):
    if not is_admin():
        flash('Admin only', 'error')
        return redirect(url_for('briefs.index'))
    designer_id = request.form.get('designer_id') or None
    db = get_db()
    brief = db.execute('SELECT * FROM briefs WHERE id = ?', (brief_id,)).fetchone()
    if brief:
        db.execute('UPDATE briefs SET assigned_to = ?, status = ? WHERE id = ?',
                   (designer_id, 'assigned' if designer_id else 'pending', brief_id))
        db.commit()
        if designer_id:
            try:
                from shared.notifications import send_notification
                send_notification(
                    int(designer_id), 'brief_assigned',
                    f'Brief assigned: {brief["name"]}',
                    message=f'You have been assigned a new lifestyle brief ({brief.get("sku_count", "?")} SKUs).',
                    link='/briefs'
                )
            except Exception:
                pass
        flash('Brief updated', 'success')
    db.close()
    return redirect(request.referrer or url_for('briefs.index'))


@briefs_bp.route('/update-brief-status/<int:brief_id>', methods=['POST'])
@tool_access_required('briefs')
def update_brief_status(brief_id):
    new_status = request.form.get('status')
    if new_status not in ['pending', 'assigned', 'in_progress', 'complete']:
        flash('Invalid status', 'error')
        return redirect(url_for('briefs.index'))
    db = get_db()
    brief = db.execute('SELECT name FROM briefs WHERE id = ?', (brief_id,)).fetchone()
    db.execute('UPDATE briefs SET status = ? WHERE id = ?', (new_status, brief_id))
    db.commit()
    db.close()
    if new_status == 'complete' and brief:
        try:
            from shared.notifications import send_to_admins
            send_to_admins(
                ntype='brief_complete',
                title=f'Brief complete: {brief["name"]}',
                message='A lifestyle brief has been marked complete.',
                link='/briefs'
            )
        except Exception:
            pass
    flash('Status updated', 'success')
    return redirect(request.referrer or url_for('briefs.index'))


@briefs_bp.route('/api/briefs/<int:brief_id>/deadline', methods=['POST'])
@tool_access_required('briefs')
def update_brief_deadline(brief_id):
    if not is_admin():
        return jsonify({'error': 'Admin only'}), 403
    data = request.get_json() or {}
    deadline = data.get('deadline', '').strip() or None
    db = get_db()
    db.execute('UPDATE briefs SET deadline = ? WHERE id = ?', (deadline, brief_id))
    db.commit()
    db.close()
    return jsonify({'success': True})


@briefs_bp.route('/update-brief-priority/<int:brief_id>', methods=['POST'])
@tool_access_required('briefs')
def update_brief_priority(brief_id):
    if not is_admin():
        flash('Admin only', 'error')
        return redirect(url_for('briefs.index'))
    new_priority = request.form.get('priority')
    if new_priority not in ['low', 'normal', 'high', 'urgent']:
        flash('Invalid priority', 'error')
        return redirect(url_for('briefs.index'))
    db = get_db()
    brief = db.execute('SELECT name FROM briefs WHERE id = ?', (brief_id,)).fetchone()
    db.execute('UPDATE briefs SET priority = ? WHERE id = ?', (new_priority, brief_id))
    db.commit()
    db.close()
    if brief:
        notify_priority_changed(brief['name'], new_priority,
                                g.user.get('display_name', g.user.get('username', '')))
    flash('Priority updated', 'success')
    return redirect(request.referrer or url_for('briefs.index'))


@briefs_bp.route('/delete-brief/<int:brief_id>', methods=['POST'])
@tool_access_required('briefs')
def delete_brief(brief_id):
    if not is_admin():
        flash('Admin only', 'error')
        return redirect(url_for('briefs.index'))
    db = get_db()
    brief = db.execute('SELECT * FROM briefs WHERE id = ?', (brief_id,)).fetchone()
    if brief:
        if brief['output_path'] and os.path.exists(brief['output_path']):
            shutil.rmtree(brief['output_path'], ignore_errors=True)
        db.execute('DELETE FROM briefs WHERE id = ?', (brief_id,))
        db.commit()
        flash('Brief deleted', 'success')
    else:
        flash('Brief not found', 'error')
    db.close()
    return redirect(url_for('briefs.index'))


@briefs_bp.route('/api/send-back-brief/<int:brief_id>', methods=['POST'])
@tool_access_required('briefs')
def send_back_brief(brief_id):
    if not is_admin():
        return jsonify({'error': 'Admin only'}), 403
    data = request.get_json() or {}
    reason = data.get('reason', '')
    db = get_db()
    brief = db.execute('SELECT * FROM briefs WHERE id = ?', (brief_id,)).fetchone()
    if not brief:
        db.close()
        return jsonify({'error': 'Brief not found'}), 404
    if brief['status'] != 'complete':
        db.close()
        return jsonify({'error': 'Brief is not complete'}), 400
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M')
    db.execute('INSERT INTO brief_amendments (brief_id, sent_back_at, sent_back_by, reason) VALUES (?, ?, ?, ?)',
               (brief_id, now_str, g.user['id'], reason))
    db.execute("UPDATE briefs SET status = 'in_progress', amendment_count = COALESCE(amendment_count, 0) + 1 WHERE id = ?",
               (brief_id,))
    db.commit()
    db.close()
    return jsonify({'success': True})


@briefs_bp.route('/api/scan-completions', methods=['POST'])
@tool_access_required('briefs')
def scan_completions():
    if not is_admin():
        flash('Admin only', 'error')
        return redirect(url_for('briefs.index'))
    db = get_db()
    briefs = db.execute(
        "SELECT id, name, sku_list FROM briefs WHERE status = 'in_progress' AND sku_list IS NOT NULL"
    ).fetchall()
    if not briefs:
        flash('No in-progress briefs to scan.', 'info')
        db.close()
        return redirect(url_for('briefs.index'))
    approved_rows = db.execute(
        "SELECT UPPER(COALESCE(confirmed_sku, detected_sku)) as sku FROM uploads WHERE status = 'approved' AND image_type = 'ls1'"
    ).fetchall()
    approved_ls1 = {r['sku'] for r in approved_rows if r['sku']}
    completed_count = 0
    for brief in briefs:
        try:
            skus = json.loads(brief['sku_list'])
        except (json.JSONDecodeError, TypeError):
            continue
        if skus and all(sku.upper() in approved_ls1 for sku in skus):
            now_str = datetime.now().strftime('%Y-%m-%d %H:%M')
            db.execute("UPDATE briefs SET status = 'complete', completed_at = ? WHERE id = ?", (now_str, brief['id']))
            db.execute("UPDATE brief_amendments SET completed_at = ? WHERE brief_id = ? AND completed_at IS NULL",
                       (now_str, brief['id']))
            db.commit()
            completed_count += 1
    db.close()
    if completed_count:
        flash(f'{completed_count} brief(s) marked complete.', 'success')
    else:
        flash(f'Scanned {len(briefs)} briefs — none fully complete yet.', 'info')
    return redirect(url_for('briefs.index'))


# ── Model Search APIs ────────────────────────────────────────────────────────

@briefs_bp.route('/api/search-models')
@tool_access_required('briefs')
def api_search_models():
    query = request.args.get('q', '').strip().lower()
    if len(query) < 2:
        return jsonify([])
    model_lookup._load_mapping()
    model_lookup._index_model_files()
    results = [n for n in model_lookup.model_files_index if query in n.lower()][:20]
    results.sort()
    return jsonify(results)


@briefs_bp.route('/api/search-mapping-models')
@tool_access_required('briefs')
def api_search_mapping_models():
    if not is_admin():
        return jsonify([])
    query = request.args.get('q', '').strip().lower()
    if len(query) < 2:
        return jsonify([])
    db = get_db()
    results = []
    seen = set()
    for row in db.execute('SELECT sku, models, model_count FROM model_mappings WHERE lower(sku) LIKE ? LIMIT 20',
                          (f'%{query}%',)).fetchall():
        if row['sku'] not in seen:
            seen.add(row['sku'])
            results.append({'type': 'sku', 'sku': row['sku'], 'models': row['models'], 'model_count': row['model_count'] or 1})
    for row in db.execute('SELECT sku, models, model_count FROM model_mappings WHERE lower(models) LIKE ? LIMIT 20',
                          (f'%{query}%',)).fetchall():
        if row['sku'] not in seen:
            seen.add(row['sku'])
            results.append({'type': 'model', 'sku': row['sku'], 'models': row['models'], 'model_count': row['model_count'] or 1})
    db.close()
    return jsonify(results[:30])


@briefs_bp.route('/api/sku-info')
@tool_access_required('briefs')
def api_sku_info():
    skus_param = request.args.get('skus', '')
    if not skus_param:
        return jsonify([])
    skus_list = [s.strip() for s in skus_param.split(',') if s.strip()]
    products = akeneo_client.get_products(skus_list)
    brand_labels = {}
    result = []
    for sku in skus_list:
        product = products.get(sku)
        category = brand_code = brand_label = ''
        if product:
            category = product['categories'][0] if product.get('categories') else ''
            brand_code = akeneo_client.get_attribute(product, 'brand') or ''
            if brand_code and brand_code not in brand_labels:
                brand_labels[brand_code] = akeneo_client.get_attribute_option_label('brand', brand_code)
            brand_label = brand_labels.get(brand_code, brand_code)
        result.append({'sku': sku, 'category': category, 'brand_code': brand_code, 'brand_label': brand_label or brand_code})
    return jsonify(result)


# ── Model Mappings ───────────────────────────────────────────────────────────

@briefs_bp.route('/model-mappings')
@tool_access_required('briefs')
def model_mappings_view():
    if not is_admin():
        flash('Admin access required', 'error')
        return redirect(url_for('briefs.index'))

    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '').strip()
    status_filter = request.args.get('status', '').strip()
    per_page = 50

    db = get_db()

    where = ['1=1']
    params = []

    if search:
        where.append('(lower(sku) LIKE ? OR lower(models) LIKE ?)')
        params.extend([f'%{search.lower()}%', f'%{search.lower()}%'])

    if status_filter == 'mapped':
        where.append('1=1')  # all rows in model_mappings are mapped
    elif status_filter == 'partial':
        where.append('model_count > 1')

    where_clause = ' WHERE ' + ' AND '.join(where)

    total = db.execute(f'SELECT COUNT(*) as c FROM model_mappings{where_clause}', params).fetchone()['c']
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))
    offset = (page - 1) * per_page

    rows = db.execute(
        f'SELECT id as mapping_id, sku, models, model_count FROM model_mappings{where_clause} ORDER BY sku ASC LIMIT ? OFFSET ?',
        params + [per_page, offset]
    ).fetchall()

    total_all = db.execute('SELECT COUNT(*) as c FROM model_mappings').fetchone()['c']
    db.close()

    return render_template('briefs/model_mappings.html',
                           rows=rows, page=page, total_pages=total_pages,
                           total=total, total_all=total_all, search=search,
                           status_filter=status_filter, per_page=per_page)


@briefs_bp.route('/api/model-mappings', methods=['POST'])
@tool_access_required('briefs')
def api_add_mapping():
    if not is_admin():
        return jsonify({'error': 'Admin only'}), 403
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data'}), 400
    sku = data.get('sku', '').strip()
    models = data.get('models', '').strip()
    if not sku or not models:
        return jsonify({'error': 'SKU and models required'}), 400
    model_count = len([m.strip() for m in models.replace(';', ',').split(',') if m.strip()])
    db = get_db()
    try:
        db.execute('INSERT INTO model_mappings (sku, models, model_count, source) VALUES (?, ?, ?, ?)',
                   (sku, models, model_count, 'manual'))
        db.commit()
        db.close()
        return jsonify({'success': True, 'message': f'Mapping added for {sku}'})
    except sqlite3.IntegrityError:
        db.close()
        return jsonify({'error': f'SKU {sku} already exists'}), 409


@briefs_bp.route('/api/model-mappings/<int:mapping_id>', methods=['PUT'])
@tool_access_required('briefs')
def api_update_mapping(mapping_id):
    if not is_admin():
        return jsonify({'error': 'Admin only'}), 403
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data'}), 400
    sku = data.get('sku', '').strip()
    models = data.get('models', '').strip()
    if not sku or not models:
        return jsonify({'error': 'SKU and models required'}), 400
    model_count = len([m.strip() for m in models.replace(';', ',').split(',') if m.strip()])
    db = get_db()
    if not db.execute('SELECT id FROM model_mappings WHERE id = ?', (mapping_id,)).fetchone():
        db.close()
        return jsonify({'error': 'Not found'}), 404
    try:
        db.execute('UPDATE model_mappings SET sku = ?, models = ?, model_count = ?, updated_at = ? WHERE id = ?',
                   (sku, models, model_count, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), mapping_id))
        db.commit()
        db.close()
        return jsonify({'success': True, 'message': f'Mapping updated for {sku}'})
    except sqlite3.IntegrityError:
        db.close()
        return jsonify({'error': 'SKU already exists on another mapping'}), 409


@briefs_bp.route('/api/model-mappings/<int:mapping_id>', methods=['DELETE'])
@tool_access_required('briefs')
def api_delete_mapping(mapping_id):
    if not is_admin():
        return jsonify({'error': 'Admin only'}), 403
    db = get_db()
    mapping = db.execute('SELECT sku FROM model_mappings WHERE id = ?', (mapping_id,)).fetchone()
    if not mapping:
        db.close()
        return jsonify({'error': 'Not found'}), 404
    db.execute('DELETE FROM model_mappings WHERE id = ?', (mapping_id,))
    db.commit()
    db.close()
    return jsonify({'success': True, 'message': f'Mapping deleted for {mapping["sku"]}'})


@briefs_bp.route('/api/model-mappings/import', methods=['POST'])
@tool_access_required('briefs')
def api_import_mappings():
    if not is_admin():
        return jsonify({'error': 'Admin only'}), 403
    csv_text = request.form.get('csv_data', '').strip()
    if not csv_text and 'csv_file' in request.files:
        f = request.files['csv_file']
        if f and f.filename:
            try:
                csv_text = f.read().decode('utf-8')
            except UnicodeDecodeError:
                return jsonify({'error': 'File must be UTF-8 encoded CSV'}), 400
    if not csv_text:
        return jsonify({'error': 'No CSV data'}), 400

    from io import StringIO as SIO
    reader = csv.DictReader(SIO(csv_text))
    if not reader.fieldnames or 'sku' not in reader.fieldnames or 'models' not in reader.fieldnames:
        return jsonify({'error': 'CSV must have "sku" and "models" columns'}), 400

    db = get_db()
    added = updated = skipped = 0
    for row in reader:
        sku = row.get('sku', '').strip()
        models = row.get('models', '').strip()
        if not sku or not models:
            skipped += 1
            continue
        model_count = len([m.strip() for m in models.replace(';', ',').split(',') if m.strip()])
        source = row.get('sources', row.get('source', 'csv_import')).strip() or 'csv_import'
        existing = db.execute('SELECT id FROM model_mappings WHERE sku = ?', (sku,)).fetchone()
        if existing:
            db.execute('UPDATE model_mappings SET models = ?, model_count = ?, source = ?, updated_at = ? WHERE id = ?',
                       (models, model_count, source, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), existing['id']))
            updated += 1
        else:
            db.execute('INSERT INTO model_mappings (sku, models, model_count, source) VALUES (?, ?, ?, ?)',
                       (sku, models, model_count, source))
            added += 1
    db.commit()
    db.close()
    model_lookup.reload()
    return jsonify({'success': True, 'added': added, 'updated': updated, 'skipped': skipped,
                    'message': f'{added} added, {updated} updated, {skipped} skipped'})


@briefs_bp.route('/api/model-mappings/sync-s3', methods=['POST'])
@tool_access_required('briefs')
def api_sync_mappings_s3():
    if not is_admin():
        return jsonify({'error': 'Admin only'}), 403
    try:
        db = get_db()
        rows = db.execute('SELECT sku, models, model_count, source FROM model_mappings ORDER BY sku').fetchall()
        db.close()
        from io import StringIO as SIO
        output = SIO()
        writer = csv.writer(output)
        writer.writerow(['sku', 'models', 'model_count', 'sources'])
        for row in rows:
            writer.writerow([row['sku'], row['models'], row['model_count'], row['source']])
        csv_content = output.getvalue()

        s3_config = CONFIG.get('s3', {})
        if s3_config.get('bucket'):
            import boto3
            s3 = boto3.client('s3',
                region_name=s3_config.get('region', 'eu-west-2'),
                aws_access_key_id=s3_config.get('access_key') or None,
                aws_secret_access_key=s3_config.get('secret_key') or None
            )
            s3.put_object(Bucket=s3_config['bucket'], Key=s3_config.get('mapping_key', 'FINAL_master_mapping.csv'),
                          Body=csv_content.encode('utf-8'), ContentType='text/csv')
        model_lookup.reload()
        return jsonify({'success': True, 'message': f'{len(rows)} mappings synced to S3'})
    except Exception as e:
        return jsonify({'error': f'Sync failed: {str(e)}'}), 500


# ── Scenes ───────────────────────────────────────────────────────────────────

@briefs_bp.route('/scenes')
@tool_access_required('briefs')
def scene_catalog():
    db = get_db()
    page = request.args.get('page', 1, type=int)
    per_page = 12
    room_filter = request.args.get('room_type', '')
    search = request.args.get('search', '').strip()

    where = []
    params = []
    if room_filter:
        where.append('s.room_type = ?')
        params.append(room_filter)
    if search:
        where.append('(LOWER(s.name) LIKE ? OR LOWER(s.tags) LIKE ? OR LOWER(s.description) LIKE ?)')
        sq = f'%{search.lower()}%'
        params.extend([sq, sq, sq])

    where_clause = (' WHERE ' + ' AND '.join(where)) if where else ''
    total = db.execute(f'SELECT COUNT(*) as c FROM scenes s{where_clause}', params).fetchone()['c']
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = min(max(1, page), total_pages)
    offset = (page - 1) * per_page

    scenes = db.execute(f'''
        SELECT s.*,
               (SELECT COUNT(*) FROM scene_files sf WHERE sf.scene_id = s.id) as file_count
        FROM scenes s
        {where_clause}
        ORDER BY s.updated_at DESC LIMIT ? OFFSET ?
    ''', params + [per_page, offset]).fetchall()

    total_scenes = db.execute('SELECT COUNT(*) as c FROM scenes').fetchone()['c']
    db.close()

    return render_template('briefs/scenes.html',
                           scenes=scenes, page=page, total_pages=total_pages,
                           total=total, total_scenes=total_scenes,
                           room_filter=room_filter, search=search, room_types=ROOM_TYPES)


@briefs_bp.route('/api/scenes', methods=['POST'])
@tool_access_required('briefs')
def api_create_scene():
    try:
        name = request.form.get('name', '').strip()
        if not name:
            return jsonify({'error': 'Scene name is required'}), 400
        description = request.form.get('description', '').strip()
        room_type = request.form.get('room_type', '').strip()
        tags = request.form.get('tags', '').strip()
        preview_file = request.files.get('preview')
        scene_files_list = request.files.getlist('scene_files')

        db = get_db()
        cursor = db.execute(
            'INSERT INTO scenes (name, description, room_type, created_by, tags) VALUES (?, ?, ?, ?, ?)',
            (name, description, room_type or None, g.user['id'], tags or None)
        )
        scene_id = cursor.lastrowid

        if preview_file and preview_file.filename:
            ext = os.path.splitext(secure_filename(preview_file.filename))[1].lower()
            preview_s3_key = f'scenes/{scene_id}/preview{ext}'
            upload_to_s3(preview_file.stream, preview_s3_key, preview_file.content_type or 'image/jpeg')
            db.execute('UPDATE scenes SET preview_s3_key = ? WHERE id = ?', (preview_s3_key, scene_id))

        for sf in scene_files_list:
            if sf and sf.filename:
                safe_name = secure_filename(sf.filename)
                s3_key = f'scenes/{scene_id}/files/{safe_name}'
                upload_to_s3(sf.stream, s3_key, sf.content_type)
                sf.stream.seek(0, 2)
                file_size = sf.stream.tell()
                sf.stream.seek(0)
                db.execute('INSERT INTO scene_files (scene_id, filename, s3_key, file_size) VALUES (?, ?, ?, ?)',
                           (scene_id, safe_name, s3_key, file_size))
        db.commit()
        db.close()
        return jsonify({'success': True, 'scene_id': scene_id, 'message': f'Scene "{name}" created'})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': f'Failed to create scene: {str(e)}'}), 500


@briefs_bp.route('/api/scenes/<int:scene_id>', methods=['PUT'])
@tool_access_required('briefs')
def api_update_scene(scene_id):
    db = get_db()
    scene = db.execute('SELECT * FROM scenes WHERE id = ?', (scene_id,)).fetchone()
    if not scene:
        db.close()
        return jsonify({'error': 'Scene not found'}), 404
    if scene['created_by'] != g.user['id'] and not is_admin():
        db.close()
        return jsonify({'error': 'Permission denied'}), 403
    data = request.get_json()
    name = data.get('name', scene['name']).strip()
    description = data.get('description', scene['description'] or '').strip()
    room_type = data.get('room_type', scene['room_type'] or '').strip()
    tags = data.get('tags', scene['tags'] or '').strip()
    db.execute('UPDATE scenes SET name = ?, description = ?, room_type = ?, tags = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?',
               (name, description, room_type or None, tags or None, scene_id))
    db.commit()
    db.close()
    return jsonify({'success': True})


@briefs_bp.route('/api/scenes/<int:scene_id>', methods=['DELETE'])
@tool_access_required('briefs')
def api_delete_scene(scene_id):
    db = get_db()
    scene = db.execute('SELECT * FROM scenes WHERE id = ?', (scene_id,)).fetchone()
    if not scene:
        db.close()
        return jsonify({'error': 'Scene not found'}), 404
    if scene['created_by'] != g.user['id'] and not is_admin():
        db.close()
        return jsonify({'error': 'Permission denied'}), 403
    try:
        for f in db.execute('SELECT s3_key FROM scene_files WHERE scene_id = ?', (scene_id,)).fetchall():
            delete_from_s3(f['s3_key'])
        if scene['preview_s3_key']:
            delete_from_s3(scene['preview_s3_key'])
    except Exception as e:
        print(f'[Scenes] S3 cleanup: {e}')
    db.execute('DELETE FROM scene_files WHERE scene_id = ?', (scene_id,))
    db.execute('DELETE FROM scenes WHERE id = ?', (scene_id,))
    db.commit()
    db.close()
    return jsonify({'success': True})


@briefs_bp.route('/api/scenes/<int:scene_id>/files', methods=['POST'])
@tool_access_required('briefs')
def api_add_scene_files(scene_id):
    db = get_db()
    scene = db.execute('SELECT * FROM scenes WHERE id = ?', (scene_id,)).fetchone()
    if not scene:
        db.close()
        return jsonify({'error': 'Scene not found'}), 404
    if scene['created_by'] != g.user['id'] and not is_admin():
        db.close()
        return jsonify({'error': 'Permission denied'}), 403
    added = 0
    for sf in request.files.getlist('scene_files'):
        if sf and sf.filename:
            safe_name = secure_filename(sf.filename)
            s3_key = f'scenes/{scene_id}/files/{safe_name}'
            upload_to_s3(sf.stream, s3_key, sf.content_type)
            sf.stream.seek(0, 2)
            file_size = sf.stream.tell()
            sf.stream.seek(0)
            db.execute('INSERT INTO scene_files (scene_id, filename, s3_key, file_size) VALUES (?, ?, ?, ?)',
                       (scene_id, safe_name, s3_key, file_size))
            added += 1
    db.execute('UPDATE scenes SET updated_at = CURRENT_TIMESTAMP WHERE id = ?', (scene_id,))
    db.commit()
    db.close()
    return jsonify({'success': True, 'added': added})


@briefs_bp.route('/api/scenes/<int:scene_id>/files/<int:file_id>', methods=['DELETE'])
@tool_access_required('briefs')
def api_delete_scene_file(scene_id, file_id):
    db = get_db()
    scene = db.execute('SELECT * FROM scenes WHERE id = ?', (scene_id,)).fetchone()
    if not scene:
        db.close()
        return jsonify({'error': 'Scene not found'}), 404
    if scene['created_by'] != g.user['id'] and not is_admin():
        db.close()
        return jsonify({'error': 'Permission denied'}), 403
    sf = db.execute('SELECT * FROM scene_files WHERE id = ? AND scene_id = ?', (file_id, scene_id)).fetchone()
    if not sf:
        db.close()
        return jsonify({'error': 'File not found'}), 404
    try:
        delete_from_s3(sf['s3_key'])
    except Exception:
        pass
    db.execute('DELETE FROM scene_files WHERE id = ?', (file_id,))
    db.execute('UPDATE scenes SET updated_at = CURRENT_TIMESTAMP WHERE id = ?', (scene_id,))
    db.commit()
    db.close()
    return jsonify({'success': True})


@briefs_bp.route('/api/scenes/<int:scene_id>/preview')
@tool_access_required('briefs')
def api_scene_preview(scene_id):
    db = get_db()
    scene = db.execute('SELECT preview_s3_key FROM scenes WHERE id = ?', (scene_id,)).fetchone()
    db.close()
    if not scene or not scene['preview_s3_key']:
        abort(404)
    try:
        url = get_s3_presigned_url(scene['preview_s3_key'])
        from flask import redirect as fredirect
        return fredirect(url)
    except Exception:
        abort(404)


@briefs_bp.route('/api/scenes/<int:scene_id>/detail')
@tool_access_required('briefs')
def api_scene_detail(scene_id):
    db = get_db()
    scene = db.execute('SELECT * FROM scenes WHERE id = ?', (scene_id,)).fetchone()
    if not scene:
        db.close()
        return jsonify({'error': 'Not found'}), 404
    files = db.execute('SELECT id, filename, file_size, uploaded_at FROM scene_files WHERE scene_id = ? ORDER BY filename',
                       (scene_id,)).fetchall()
    db.close()
    return jsonify({
        'id': scene['id'], 'name': scene['name'], 'description': scene['description'],
        'room_type': scene['room_type'], 'tags': scene['tags'],
        'created_at': scene['created_at'], 'updated_at': scene['updated_at'],
        'has_preview': bool(scene['preview_s3_key']),
        'files': [{'id': f['id'], 'filename': f['filename'], 'file_size': f['file_size'],
                   'uploaded_at': f['uploaded_at']} for f in files]
    })


@briefs_bp.route('/api/scenes/<int:scene_id>/download/<int:file_id>')
@tool_access_required('briefs')
def api_download_scene_file(scene_id, file_id):
    db = get_db()
    sf = db.execute('SELECT * FROM scene_files WHERE id = ? AND scene_id = ?', (file_id, scene_id)).fetchone()
    db.close()
    if not sf:
        abort(404)
    try:
        from flask import redirect as fredirect
        return fredirect(get_s3_presigned_url(sf['s3_key']))
    except Exception:
        abort(404)


@briefs_bp.route('/api/scenes/search')
@tool_access_required('briefs')
def api_search_scenes():
    q = request.args.get('q', '').strip().lower()
    db = get_db()
    where = []
    params = []
    if q and len(q) >= 2:
        where.append('(LOWER(s.name) LIKE ? OR LOWER(s.tags) LIKE ?)')
        params.extend([f'%{q}%', f'%{q}%'])
    where_clause = (' WHERE ' + ' AND '.join(where)) if where else ''
    scenes = db.execute(f'''
        SELECT s.id, s.name, s.room_type, s.tags, s.preview_s3_key
        FROM scenes s {where_clause}
        ORDER BY s.updated_at DESC LIMIT 20
    ''', params).fetchall()
    db.close()
    return jsonify([{
        'id': s['id'], 'name': s['name'],
        'room_type': s['room_type'] or '', 'tags': s['tags'] or '',
        'has_preview': bool(s['preview_s3_key'])
    } for s in scenes])


# ── Uploads ──────────────────────────────────────────────────────────────────

@briefs_bp.route('/uploads')
@tool_access_required('briefs')
def uploads_page():
    db = get_db()
    user = g.user
    uploads = db.execute('''
        SELECT u.* FROM uploads u WHERE u.uploaded_by = ? ORDER BY u.created_at DESC
    ''', (user['id'],)).fetchall()
    if user['role'] == 'admin':
        briefs = db.execute("SELECT id, name, sku_count FROM briefs ORDER BY created_at DESC").fetchall()
    else:
        briefs = db.execute("SELECT id, name, sku_count FROM briefs WHERE user_id = ? ORDER BY created_at DESC",
                            (user['id'],)).fetchall()
    db.close()
    return render_template('briefs/uploads.html', uploads=uploads, briefs=briefs)


@briefs_bp.route('/uploads/review')
@tool_access_required('briefs')
def uploads_review():
    if not is_admin():
        flash('Admin access required', 'error')
        return redirect(url_for('briefs.uploads_page'))
    db = get_db()
    status_filter = request.args.get('status', 'pending')
    uploads = db.execute('''
        SELECT u.*, b.name as brief_name
        FROM uploads u
        LEFT JOIN briefs b ON u.brief_id = b.id
        WHERE u.status = ?
        ORDER BY u.created_at DESC
    ''', (status_filter,)).fetchall()
    counts = {}
    for s in ['pending', 'approved', 'rejected']:
        counts[s] = db.execute('SELECT COUNT(*) as c FROM uploads WHERE status = ?', (s,)).fetchone()['c']
    db.close()
    return render_template('briefs/uploads_review.html', uploads=uploads,
                           status_filter=status_filter, counts=counts)


@briefs_bp.route('/api/uploads', methods=['POST'])
@tool_access_required('briefs')
def api_upload_files():
    files = request.files.getlist('files')
    if not files or not files[0].filename:
        return jsonify({'error': 'No files selected'}), 400
    brief_id = request.form.get('brief_id', '').strip() or None
    sku_override = request.form.get('sku_override', '').strip() or None
    allowed_ext = {'.jpg', '.jpeg', '.png', '.tiff', '.tif'}
    db = get_db()
    uploaded = []
    for f in files:
        ext = os.path.splitext(f.filename)[1].lower()
        if ext not in allowed_ext:
            continue
        safe_name = secure_filename(f.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        s3_key = f'uploads/pending/{g.user["id"]}/{timestamp}_{safe_name}'
        upload_to_s3(f.stream, s3_key, f.content_type)
        detected_sku, image_type = detect_sku_from_filename(safe_name)
        if sku_override:
            detected_sku = sku_override
        db.execute('''INSERT INTO uploads (uploaded_by, original_filename, s3_key, detected_sku, image_type, brief_id, status)
                      VALUES (?, ?, ?, ?, ?, ?, 'pending')''',
                   (g.user['id'], safe_name, s3_key, detected_sku, image_type, brief_id))
        uploaded.append(safe_name)
    db.commit()
    db.close()
    if uploaded:
        return jsonify({'success': True, 'count': len(uploaded), 'files': uploaded})
    return jsonify({'error': 'No valid image files. Allowed: JPG, PNG, TIFF'}), 400


@briefs_bp.route('/api/uploads/<int:upload_id>/preview')
@tool_access_required('briefs')
def api_upload_preview(upload_id):
    db = get_db()
    upload = db.execute('SELECT * FROM uploads WHERE id = ?', (upload_id,)).fetchone()
    db.close()
    if not upload:
        abort(404)
    if not is_admin() and upload['uploaded_by'] != g.user['id']:
        abort(403)
    try:
        s3 = _get_s3_client()
        s3_obj = s3.get_object(Bucket=CONFIG['s3']['bucket'], Key=upload['s3_key'])
        image_data = s3_obj['Body'].read()
        content_type = s3_obj.get('ContentType', 'image/jpeg')
        return Response(image_data, mimetype=content_type, headers={'Cache-Control': 'private, max-age=300'})
    except Exception:
        abort(404)


@briefs_bp.route('/api/cutout/<sku>')
@tool_access_required('briefs')
def api_cutout_preview(sku):
    try:
        product = akeneo_client.get_product(sku)
        if not product:
            abort(404)
        cutout_url = akeneo_client.get_asset_url(product, 'cutout_1')
        if not cutout_url:
            abort(404)
        resp = requests.get(cutout_url, timeout=10)
        if resp.status_code != 200:
            abort(404)
        return Response(resp.content, mimetype=resp.headers.get('Content-Type', 'image/jpeg'),
                        headers={'Cache-Control': 'private, max-age=3600'})
    except Exception:
        abort(404)


@briefs_bp.route('/api/uploads/<int:upload_id>/approve', methods=['POST'])
@tool_access_required('briefs')
def api_approve_upload(upload_id):
    if not is_admin():
        return jsonify({'error': 'Admin only'}), 403
    db = get_db()
    upload = db.execute('SELECT * FROM uploads WHERE id = ?', (upload_id,)).fetchone()
    if not upload:
        db.close()
        return jsonify({'error': 'Upload not found'}), 404
    if upload['status'] != 'pending':
        db.close()
        return jsonify({'error': 'Upload already reviewed'}), 400
    data = request.get_json() or {}
    confirmed_sku = data.get('sku', upload['detected_sku'] or '').strip()
    image_type = data.get('image_type', upload['image_type'] or '').strip()
    if not confirmed_sku:
        db.close()
        return jsonify({'error': 'SKU is required'}), 400
    if not image_type:
        db.close()
        return jsonify({'error': 'Image type (ls1/ls2/ls3/sp1) is required'}), 400
    if confirmed_sku.lower().endswith(f'_{image_type}'):
        confirmed_sku = confirmed_sku[:-(len(image_type) + 1)]
    try:
        s3 = _get_s3_client()
        s3_obj = s3.get_object(Bucket=CONFIG['s3']['bucket'], Key=upload['s3_key'])
        file_data = s3_obj['Body'].read()
        ext = os.path.splitext(upload['original_filename'])[1].lower()
        if ext not in ['.jpg', '.jpeg', '.png']:
            ext = '.jpg'
        scaleflex_filename = f'{confirmed_sku}_{image_type}{ext}'
        scaleflex_client.upload_file(io.BytesIO(file_data), scaleflex_filename)
    except Exception as e:
        db.close()
        return jsonify({'error': f'Scaleflex upload failed: {str(e)}'}), 500
    try:
        delete_from_s3(upload['s3_key'])
    except Exception:
        pass
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    db.execute('''UPDATE uploads SET status = 'approved', confirmed_sku = ?, image_type = ?,
                  reviewer_id = ?, reviewed_at = ? WHERE id = ?''',
               (confirmed_sku, image_type, g.user['id'], now_str, upload_id))
    db.commit()
    brief_completed = False
    brief_id = upload['brief_id']
    if brief_id:
        brief = db.execute(
            "SELECT id, name, status, sku_list, completed_at FROM briefs WHERE id = ? AND status IN ('assigned', 'in_progress')",
            (brief_id,)
        ).fetchone()
        if brief and brief['sku_list']:
            try:
                skus = json.loads(brief['sku_list'])
            except (json.JSONDecodeError, TypeError):
                skus = []
            if skus:
                approved_ls1 = {r['sku'] for r in db.execute(
                    "SELECT UPPER(COALESCE(confirmed_sku, detected_sku)) as sku FROM uploads WHERE status = 'approved' AND image_type = 'ls1'"
                ).fetchall() if r['sku']}
                if all(sku.upper() in approved_ls1 for sku in skus):
                    db.execute("UPDATE briefs SET status = 'complete', completed_at = ? WHERE id = ?", (now_str, brief['id']))
                    db.commit()
                    brief_completed = True
    db.close()
    result = {'success': True, 'filename': scaleflex_filename}
    if brief_completed:
        result['brief_completed'] = True
    return jsonify(result)


@briefs_bp.route('/api/uploads/<int:upload_id>/reject', methods=['POST'])
@tool_access_required('briefs')
def api_reject_upload(upload_id):
    if not is_admin():
        return jsonify({'error': 'Admin only'}), 403
    db = get_db()
    upload = db.execute('SELECT * FROM uploads WHERE id = ?', (upload_id,)).fetchone()
    if not upload:
        db.close()
        return jsonify({'error': 'Upload not found'}), 404
    if upload['status'] != 'pending':
        db.close()
        return jsonify({'error': 'Upload already reviewed'}), 400
    data = request.get_json() or {}
    comment = data.get('comment', '').strip()
    if not comment:
        db.close()
        return jsonify({'error': 'Please provide a reason for rejection'}), 400
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    db.execute('''UPDATE uploads SET status = 'rejected', review_comment = ?,
                  reviewer_id = ?, reviewed_at = ? WHERE id = ?''',
               (comment, g.user['id'], now_str, upload_id))
    db.commit()
    db.close()
    return jsonify({'success': True})


@briefs_bp.route('/api/uploads/<int:upload_id>/resubmit', methods=['POST'])
@tool_access_required('briefs')
def api_resubmit_upload(upload_id):
    db = get_db()
    upload = db.execute('SELECT * FROM uploads WHERE id = ?', (upload_id,)).fetchone()
    if not upload:
        db.close()
        return jsonify({'error': 'Upload not found'}), 404
    if upload['uploaded_by'] != g.user['id']:
        db.close()
        return jsonify({'error': 'Permission denied'}), 403
    if upload['status'] != 'rejected':
        db.close()
        return jsonify({'error': 'Only rejected uploads can be resubmitted'}), 400
    f = request.files.get('file')
    if not f or not f.filename:
        db.close()
        return jsonify({'error': 'No file provided'}), 400
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in {'.jpg', '.jpeg', '.png', '.tiff', '.tif'}:
        db.close()
        return jsonify({'error': 'Invalid file type'}), 400
    safe_name = secure_filename(f.filename)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    new_s3_key = f'uploads/pending/{g.user["id"]}/{timestamp}_{safe_name}'
    try:
        delete_from_s3(upload['s3_key'])
    except Exception:
        pass
    upload_to_s3(f.stream, new_s3_key, f.content_type)
    detected_sku, image_type = detect_sku_from_filename(safe_name)
    db.execute('''UPDATE uploads SET status = 'pending', original_filename = ?, s3_key = ?,
                  detected_sku = ?, image_type = ?, review_comment = NULL,
                  reviewer_id = NULL, reviewed_at = NULL WHERE id = ?''',
               (safe_name, new_s3_key, detected_sku, image_type, upload_id))
    db.commit()
    db.close()
    return jsonify({'success': True})


@briefs_bp.route('/api/uploads/<int:upload_id>/delete', methods=['POST'])
@tool_access_required('briefs')
def api_delete_upload(upload_id):
    if not is_admin():
        return jsonify({'error': 'Admin only'}), 403
    db = get_db()
    upload = db.execute('SELECT * FROM uploads WHERE id = ?', (upload_id,)).fetchone()
    if not upload:
        db.close()
        return jsonify({'error': 'Upload not found'}), 404
    try:
        delete_from_s3(upload['s3_key'])
    except Exception:
        pass
    db.execute('DELETE FROM uploads WHERE id = ?', (upload_id,))
    db.commit()
    db.close()
    return jsonify({'success': True})


@briefs_bp.route('/api/uploads/bulk', methods=['POST'])
@tool_access_required('briefs')
def api_bulk_uploads():
    if not is_admin():
        return jsonify({'error': 'Admin only'}), 403
    data = request.get_json() or {}
    ids = data.get('ids', [])
    action = data.get('action', '')
    comment = data.get('comment', '').strip()
    if not ids:
        return jsonify({'error': 'No uploads selected'}), 400
    if action not in ['delete', 'approve', 'reject']:
        return jsonify({'error': 'Invalid action'}), 400

    db = get_db()
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    success = 0
    errors = []

    for uid in ids:
        upload = db.execute('SELECT * FROM uploads WHERE id = ?', (uid,)).fetchone()
        if not upload:
            continue
        if action == 'delete':
            try:
                delete_from_s3(upload['s3_key'])
            except Exception:
                pass
            db.execute('DELETE FROM uploads WHERE id = ?', (uid,))
            success += 1
        elif action == 'approve':
            if upload['status'] != 'pending':
                continue
            sku = upload['detected_sku'] or ''
            img_type = upload['image_type'] or ''
            if not sku or not img_type:
                errors.append(f'{upload["original_filename"]}: missing SKU or type')
                continue
            if sku.lower().endswith(f'_{img_type}'):
                sku = sku[:-(len(img_type) + 1)]
            try:
                s3 = _get_s3_client()
                s3_obj = s3.get_object(Bucket=CONFIG['s3']['bucket'], Key=upload['s3_key'])
                file_data = s3_obj['Body'].read()
                ext = os.path.splitext(upload['original_filename'])[1].lower()
                if ext not in ['.jpg', '.jpeg', '.png']:
                    ext = '.jpg'
                sf_name = f'{sku}_{img_type}{ext}'
                scaleflex_client.upload_file(io.BytesIO(file_data), sf_name)
                try:
                    delete_from_s3(upload['s3_key'])
                except Exception:
                    pass
                db.execute('''UPDATE uploads SET status = 'approved', confirmed_sku = ?, image_type = ?,
                              reviewer_id = ?, reviewed_at = ? WHERE id = ?''',
                           (sku, img_type, g.user['id'], now_str, uid))
                success += 1
            except Exception as e:
                errors.append(f'{upload["original_filename"]}: {str(e)[:50]}')
        elif action == 'reject':
            if upload['status'] != 'pending':
                continue
            if not comment:
                return jsonify({'error': 'Comment required for rejection'}), 400
            db.execute('''UPDATE uploads SET status = 'rejected', review_comment = ?,
                          reviewer_id = ?, reviewed_at = ? WHERE id = ?''',
                       (comment, g.user['id'], now_str, uid))
            success += 1

    db.commit()
    db.close()
    result = {'success': True, 'count': success}
    if errors:
        result['errors'] = errors
    return jsonify(result)
